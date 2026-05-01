from __future__ import annotations

import logging
import re
import warnings
from datetime import date, datetime, time
from pathlib import Path
from typing import (
    TYPE_CHECKING,
    BinaryIO,
    Iterable,
    Iterator,
    Literal,
    NamedTuple,
    Optional,
    TypeAlias,
    Union,
    overload,
)

if TYPE_CHECKING:
    from mireport.taxonomy import Concept

from dateutil.parser import parse as parse_datetime
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell, MergedCell, ReadOnlyCell
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from mireport.conversionresults import ConversionResultsBuilder, MessageType, Severity
from mireport.exceptions import OpenPyXlRelatedException
from mireport.typealiases import DecimalPlaces
from mireport.xlsx_template_reader._bindings import (
    CellAndXBRLMetadataHolder,
    CellRangeMetadata,
)

L = logging.getLogger(__name__)

CellType: TypeAlias = ReadOnlyCell | MergedCell | Cell
CellValueType: TypeAlias = bool | float | int | str | datetime | date | time | None

EXCEL_PLACEHOLDER_VALUE = "#VALUE!"
EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE = frozenset({"-", EXCEL_PLACEHOLDER_VALUE})


def conceptsToText(concepts: Iterable[Concept]) -> str:
    return ", ".join(sorted(str(c.qname) for c in concepts))


class NamedRangeException(OpenPyXlRelatedException):
    """Exception raised when a named range is broken in the workbook."""

    def __init__(self, message: str, defined_name: DefinedName) -> None:
        self.message = message
        self.defined_name = defined_name
        super().__init__(message, str(defined_name))

    def __str__(self) -> str:
        details = (
            f"Details:\n"
            f"  Name: {self.defined_name.name}\n"
            f"  Refers to: {self.defined_name.attr_text}\n"
        )
        return f"{self.message} {details}"


def checkExcelFilePath(path: Path) -> None:
    if not path.is_file():
        raise FileNotFoundError(f'"{path}" is not a file.')
    elif path.suffix != ".xlsx":
        raise Exception(f'"{path}" is not a supported (.xlsx) Excel file')


def loadExcelFromPathOrFileLike(
    pathOrFile: Path | BinaryIO, read_only: bool = False
) -> Workbook:
    # We can safely suppress these warnings as our use-case is **just**
    # extracting data from the Excel file.
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=".*? extension is not supported and will be removed",
            category=UserWarning,
            module=r"openpyxl\.worksheet\._reader",
        )
        wb = load_workbook(
            filename=pathOrFile, read_only=read_only, data_only=True, rich_text=True
        )
    return wb


def excelCellRef(worksheet: Worksheet, cell: CellType) -> str:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    ref = f"{quote_sheetname(worksheet.title)}!{absolute_coordinate(cell.coordinate)}"
    return ref


def excelCellRangeRef(worksheet: Worksheet, cellRange: CellRange) -> str:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    ref = f"{quote_sheetname(worksheet.title)}!{absolute_coordinate(cellRange.coord)}"
    return ref


def excelCellOrCellRangeRef(
    worksheet: Worksheet, cellRange: CellRange, cell: CellType | None
) -> str:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    if cell is not None:
        return excelCellRef(worksheet, cell)
    elif cellRange is not None:
        return excelCellRangeRef(worksheet, cellRange)
    else:
        return None


def excelDefinedNameRef(
    definedName: Optional[DefinedName], cell: Optional[CellType] = None
) -> Optional[str]:
    """Make an Excel cell reference such as 'Example sheet'!$A$5"""
    if definedName is None:
        return None

    destinations = list(definedName.destinations)
    match len(destinations):
        case 1:
            sheet_name, cell_range = destinations[0]
            if cell is not None:
                coord = cell.coordinate
            else:
                coord = cell_range
            ref = f"{quote_sheetname(sheet_name)}!{absolute_coordinate(coord)}"
            return ref
        case _:
            return None


def getNamedRanges(
    wb: Workbook,
) -> tuple[dict[str, list[CellValueType]], list[NamedRangeException]]:
    data = {}
    errors = []
    for dn in list(wb.defined_names.values()):
        if not dn.name:
            errors.append(
                NamedRangeException("Named range exists but has no name.", dn)
            )
            continue

        if not dn.destinations:
            errors.append(
                NamedRangeException("Named range has no destination specified.", dn)
            )
            continue

        sheet_name, cell_range = list(dn.destinations)[0]

        if sheet_name not in wb:
            errors.append(
                NamedRangeException(
                    "Named range refers to a worksheet that is not in the workbook.", dn
                )
            )
            continue

        ws = wb[sheet_name]

        if not cell_range:
            errors.append(
                NamedRangeException("Named range has no cell range specified.", dn)
            )
            continue

        cr = CellRange(cell_range)

        if (
            cr.min_col is None
            or cr.min_row is None
            or cr.max_col is None
            or cr.max_row is None
        ):
            errors.append(
                NamedRangeException(
                    f"Named range cell range bounds expected to be int but actually None {cr=}.",
                    dn,
                )
            )
            continue

        width: int = cr.max_col - cr.min_col
        height: int = cr.max_row - cr.min_row
        if width < 0 or height < 0:
            errors.append(
                NamedRangeException(
                    f"Named range has negative cell range {width=} {height=}.", dn
                )
            )
            continue

        if not width and not height:
            # a single (width=0, height=0) cell range ... so the OpenPyXL API returns it directly.
            cell = ws[cell_range]
            values = [cell.value]
        else:
            values = []
            for row in ws[cell_range]:
                values.extend([c.value for c in row])
        data[dn.name] = values

    return data, errors


def get_decimal_places(cell: CellType) -> DecimalPlaces:
    """
    Returns the number of decimal places in the cell's number format. For
    example, a format of '0.00' would return 2.

    If no decimal places are specified, return Literal['INF'], meaning infinite
    precision, include all digits in display.
    """
    number_format = cell.number_format

    # Match typical decimal number formats like '0.00', '#,##0.000', etc.
    match = re.search(r"\.(0+)", number_format)
    if match:
        return len(match.group(1))

    # Handle cases like percentage formats '0.0%' or '0.000%'
    match_percent = re.search(r"\.(0+)%", number_format)
    if match_percent:
        return len(match_percent.group(1))

    # Catch general cases like scientific notation '0.00E+00'
    match_sci = re.search(r"\.(0+)[eE]", number_format)
    if match_sci:
        return len(match_sci.group(1))

    return "INF"


@overload
def getCellRangeIterator(
    ws: Worksheet,
    cr: CellRange,
    row_start: Optional[int] = None,
    col_start: Optional[int] = None,
    group_by_row: Literal[False] = False,
) -> Iterator[tuple[int, int, CellType]]: ...


@overload
def getCellRangeIterator(
    ws: Worksheet,
    cr: CellRange,
    row_start: Optional[int] = None,
    col_start: Optional[int] = None,
    group_by_row: Literal[True] = True,
) -> Iterator[tuple[int, tuple[CellType, ...]]]: ...


def getCellRangeIterator(
    ws: Worksheet,
    cr: CellRange,
    row_start: Optional[int] = None,
    col_start: Optional[int] = None,
    group_by_row: bool = False,
) -> Iterator[Union[tuple[int, int, CellType], tuple[int, tuple[CellType, ...]]]]:
    """Iterates over cells in the given range, supporting both standard and row-grouped modes."""

    if cr.min_row is None or cr.min_col is None:
        raise OpenPyXlRelatedException(
            f"Cell range bounds expected to be int but actually None {cr=}"
        )
    actual_row_start: int = cr.min_row
    if row_start is not None:
        actual_row_start = row_start

    actual_col_start: int = cr.min_col
    if col_start is not None:
        actual_col_start = col_start

    for rnum, row in enumerate(
        ws.iter_rows(
            min_row=cr.min_row,
            min_col=cr.min_col,
            max_row=cr.max_row,
            max_col=cr.max_col,
        ),
        start=actual_row_start,
    ):
        if group_by_row:
            yield rnum, row  # Yield row number and tuple of cells
        else:
            for cnum, cell in enumerate(row, start=actual_col_start):
                yield rnum, cnum, cell  # Yield row number, column number, and cell


class CellRangeDimensions(NamedTuple):
    width: int
    height: int
    cellsAccessed: set[tuple[str, int, int]]
    cellsPopulated: set[tuple[str, int, int]]

    @property
    def countAccessed(self) -> int:
        return len(self.cellsAccessed)

    @property
    def countPopulated(self) -> int:
        return len(self.cellsPopulated)


def getDateFromValue(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    elif isinstance(value, date):
        return value
    elif isinstance(value, str):
        if "-" in value:
            return date.fromisoformat(value)
        elif "/" in value:
            return parse_datetime(value, yearfirst=False, dayfirst=True).date()
        raise ValueError(f"Unsupported date string: '{value}'")
    else:
        raise TypeError(f"Unsupported type for date conversion: {type(value).__name__}")


def getEffectiveCellRangeDimensions(
    ws: Worksheet, cell_range: CellRange
) -> CellRangeDimensions:
    cols_not_empty: set[int] = set()
    cols_empty: set[int] = set()
    populated_rows: set[int] = set()
    populatedCellCount: set[tuple[str, int, int]] = set()
    cellCount: set[tuple[str, int, int]] = set()

    last_rnum = None
    empty_row = True
    sheetName = ws.title
    for rnum, cnum, cell in getCellRangeIterator(ws, cell_range):
        cellCount.add((sheetName, rnum, cnum))
        if last_rnum is None:
            last_rnum = rnum

        if rnum != last_rnum:
            if not empty_row:
                populated_rows.add(last_rnum)
            last_rnum = rnum
            empty_row = True

        if cell.value is not None:
            populatedCellCount.add((sheetName, rnum, cnum))
            empty_row = False
            cols_not_empty.add(cnum)
        else:
            cols_empty.add(cnum)
    else:
        if not empty_row:
            populated_rows.add(rnum)

    definitely_empty_cols = cols_empty - cols_not_empty
    total_cols = len(cols_not_empty.union(cols_empty))
    width = max(1, total_cols - len(definitely_empty_cols))
    height = max(1, len(populated_rows))
    return CellRangeDimensions(
        width=width,
        height=height,
        cellsAccessed=cellCount,
        cellsPopulated=populatedCellCount,
    )


class WorkbookReader:
    """Ergonomic cell-level access to an openpyxl Workbook.

    Carries the three shared context items (workbook, unused-name set,
    results builder) so call sites don't repeat them on every helper call.
    """

    def __init__(
        self,
        workbook: Workbook,
        unused: set[DefinedName],
        results: ConversionResultsBuilder,
    ) -> None:
        self._workbook = workbook
        self._unused = unused
        self._results = results

    def close(self) -> None:
        self._workbook.close()

    def getDefinedName(self, name: str) -> Optional[DefinedName]:
        return self._workbook.defined_names.get(name)

    def _getCellRange(self, dn: DefinedName) -> Optional[CellRangeMetadata]:
        all_destinations = list(dn.destinations)
        match len(all_destinations):
            case 0:
                self._results.addMessage(
                    f"Named range {dn.name} has no destinations specified. Ignoring.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                )
                return None
            case 1:
                pass
            case _:
                self._results.addMessage(
                    f"Table {dn.name} has multiple destinations. Ignoring table.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                )
                return None
        sheetName, cell_range = all_destinations[0]
        if not sheetName or not cell_range:
            self._results.addMessage(
                f"Named range {dn.name} has damaged cell reference {sheetName=} {cell_range=}",
                Severity.ERROR,
                MessageType.ExcelParsing,
            )
            return None
        try:
            ws = self._workbook[sheetName]
            cr = CellRange(cell_range)
        except Exception as e:
            L.exception("OpenPyXL is sad.", exc_info=e)
            return None
        dims = getEffectiveCellRangeDimensions(ws, cr)
        self._results.addCellQueries(dims.cellsAccessed)
        self._results.addCellsWithData(dims.cellsPopulated)
        return CellRangeMetadata(
            dn,
            ws,
            cr,
            effectiveHeight=dims.height,
            effectiveWidth=dims.width,
            cellsPopulated=len(dims.cellsPopulated),
        )

    def getSingleCell(
        self,
        definedName: DefinedName | str | CellAndXBRLMetadataHolder | CellRangeMetadata,
        *,
        row: int = -1,
        column: int = -1,
    ) -> Optional[CellType]:
        if isinstance(definedName, str):
            definedName = self._workbook.defined_names.get(definedName)
            if definedName is None:
                return None

        if isinstance(definedName, DefinedName):
            crh = self._getCellRange(definedName)
            if crh is None:
                return None
            self._unused.discard(definedName)
            stuff = crh
        elif isinstance(definedName, (CellAndXBRLMetadataHolder, CellRangeMetadata)):
            stuff = definedName
        else:
            return None

        self._unused.discard(stuff.definedName)

        cr = stuff.cellRange
        ws = stuff.worksheet

        if not all(
            x is not None for x in (cr.min_row, cr.max_row, cr.min_col, cr.max_col)
        ):
            self._results.addMessage(
                f"Named range {stuff.definedName.name} has an invalid cell range {cr.bounds}.",
                Severity.ERROR,
                MessageType.DevInfo,
                excel_reference=excelDefinedNameRef(stuff.definedName),
            )
            return None

        if cr.min_row == cr.max_row:
            row = cr.min_row
        if cr.min_col == cr.max_col:
            column = cr.min_col

        if row == -1:
            row = cr.min_row
        if column == -1:
            column = cr.min_col

        if not (cr.min_row <= row <= cr.max_row):
            self._results.addMessage(
                f"Row {row} has not been specified correctly.",
                Severity.ERROR,
                MessageType.DevInfo,
                excel_reference=excelCellRangeRef(ws, cr),
            )
            row = cr.min_row
        if not (cr.min_col <= column <= cr.max_col):
            self._results.addMessage(
                f"Column {column} has not been specified correctly.",
                Severity.ERROR,
                MessageType.DevInfo,
                excel_reference=excelCellRangeRef(ws, cr),
            )
            column = cr.min_col

        rows = list(
            ws.iter_rows(min_row=row, max_row=row, min_col=column, max_col=column)
        )
        match len(rows):
            case 0:
                return None
            case 1:
                cells = rows[0]
            case _:
                return None

        match len(cells):
            case 0:
                cell = None
                self._results.addMessage(
                    "No cells found in row of this named range.",
                    Severity.ERROR,
                    MessageType.DevInfo,
                    excel_reference=excelCellRangeRef(ws, cr),
                )
            case 1:
                cell = cells[0]
            case _:
                cell = None
                self._results.addMessage(
                    f"More than one cell found in range but only expected one cell. {cells}",
                    Severity.ERROR,
                    MessageType.DevInfo,
                    excel_reference=excelCellRangeRef(ws, cr),
                )

        if cell is None or cell.value is None:
            return None

        if cell.value == EXCEL_PLACEHOLDER_VALUE:
            self._results.addMessage(
                f"Excel cell has an invalid stored value {EXCEL_PLACEHOLDER_VALUE}. Please check the Excel formula for this specific cell.",
                Severity.ERROR,
                MessageType.ExcelParsing,
                excel_reference=excelCellRef(ws, cell),
            )
            return None
        return cell

    def getSingleValue(
        self,
        definedName: DefinedName | str,
        *,
        row: int = -1,
        column: int = -1,
    ) -> CellValueType:
        if (
            cell := self.getSingleCell(definedName, row=row, column=column)
        ) is not None:
            value = cell.value
            if not isinstance(value, CellValueType):
                value = str(value)
            return value
        return None

    def getSingleStringValue(
        self,
        definedName: DefinedName | str,
        *,
        row: int = -1,
        column: int = -1,
        fallbackValue: str = "",
    ) -> str:
        value = self.getSingleValue(definedName, row=row, column=column)
        return str(value) if value is not None else str(fallbackValue)

    def getSingleDateValue(self, definedName: DefinedName | str) -> date:
        value = self.getSingleValue(definedName)
        return getDateFromValue(value)
