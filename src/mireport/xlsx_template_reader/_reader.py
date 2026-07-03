from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange

from mireport.conversionresults import ConversionResultsBuilder, MessageType
from mireport.xlsx_template_reader._constants import (
    EXCEL_PLACEHOLDER_VALUE,
    EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE,
    IGNORED_DEFINED_NAME_PREFIXES,
    CellType,
    CellValueType,
)
from mireport.xlsx_template_reader._messages import Messenger
from mireport.xlsx_template_reader._ranges import (
    CellRangeMetadata,
    getEffectiveCellRangeDimensions,
)
from mireport.xlsx_template_reader.util import (
    excelDefinedNameRef,
    getDateFromValue,
)

L = logging.getLogger(__name__)


@dataclass(frozen=True, slots=True)
class CellValue:
    """A single cell value from a named range, with typed accessors."""

    raw: CellValueType

    @classmethod
    def fromCell(cls, cell: Optional[CellType]) -> CellValue:
        """Wrap a cell's value, stringifying rich objects (e.g. rich text) that
        aren't plain cell value types. A missing cell yields a blank CellValue."""
        if cell is None:
            return cls(None)
        v = cell.value
        if not isinstance(v, CellValueType):
            v = str(v)
        return cls(v)

    @property
    def hasValue(self) -> bool:
        """True if the cell holds any value at all (even a placeholder like '-').

        Use isBlank for the wider "nothing usable here" check.
        """
        return self.raw is not None

    @property
    def isBlank(self) -> bool:
        """True for empty cells and for Excel placeholder values ('-', '#VALUE!')."""
        if self.raw is None:
            return True
        if isinstance(self.raw, str):
            stripped = self.raw.strip()
            return not stripped or stripped in EXCEL_VALUES_TO_BE_TREATED_AS_NONE_VALUE
        return False

    def as_str(self, *, fallback: str = "") -> str:
        return str(self.raw) if self.raw is not None else fallback

    def as_str_stripped(self, *, fallback: str = "") -> str:
        """as_str() with surrounding whitespace removed. The fallback is for
        missing values only; a whitespace-only cell strips to ''."""
        return str(self.raw).strip() if self.raw is not None else fallback

    def as_date(self) -> date:
        """Interpret the value as a date. Raises ValueError/TypeError if it isn't one."""
        return getDateFromValue(self.raw)


class WorkbookReader:
    """Ergonomic cell-level access to an openpyxl Workbook.

    Carries the workbook and results builder, with internal tracking of
    unused named ranges.
    """

    def __init__(
        self,
        workbook: Workbook,
        results: ConversionResultsBuilder,
    ) -> None:
        self._workbook = workbook
        self._unused: set[DefinedName] = {
            dn
            for dn in workbook.defined_names.values()
            if (name := dn.name) and not name.startswith(IGNORED_DEFINED_NAME_PREFIXES)
        }
        self._results = results
        self._msg = Messenger(results)

    def close(self) -> None:
        self._workbook.close()

    def getDefinedName(self, name: str) -> Optional[DefinedName]:
        return self._workbook.defined_names.get(name)

    @property
    def unused_defined_names(self) -> frozenset[DefinedName]:
        return frozenset(self._unused)

    def markUsed(self, dn: DefinedName) -> None:
        """Record that this defined name has been consumed by the conversion."""
        self._unused.discard(dn)

    def peekRange(self, dn: DefinedName) -> Optional[CellRangeMetadata]:
        """Resolve a defined name to its cell range without marking it used.

        Emits an error message and returns None when the defined name is
        damaged (unreadable, zero or multiple destinations, broken reference).
        """
        try:
            all_destinations = list(dn.destinations)
        except AttributeError:
            self._msg.error(
                f"Named range {dn.name} has an unreadable destination: {dn.attr_text!r}. \nSomething has modified the digital template's structure. \nPlease try a fresh copy of the template and check that it has not been modified in unsupported ways.",
                MessageType.DevInfo,
            )
            L.debug(
                f"OpenPyXL error processing named range definition {dn.name=} {dn.attr_text=!r}.",
                exc_info=True,
            )
            return None
        match len(all_destinations):
            case 0:
                self._msg.error(
                    f"Named range {dn.name} has no destinations specified. Ignoring.",
                    MessageType.DevInfo,
                )
                return None
            case 1:
                pass
            case _:
                self._msg.error(
                    f"Table {dn.name} has multiple destinations. Ignoring table.",
                    MessageType.DevInfo,
                )
                return None
        sheetName, cell_range = all_destinations[0]
        if not sheetName or not cell_range:
            self._msg.error(
                f"Named range {dn.name} has damaged cell reference {sheetName=} {cell_range=}",
                MessageType.ExcelParsing,
            )
            return None
        try:
            ws = self._workbook[sheetName]
            cr = CellRange(cell_range)
        except Exception as e:
            L.debug(
                f"OpenPyXL error processing cell range. {dn.name=} {sheetName=} {cell_range=}",
                exc_info=e,
            )
            return None
        dims = getEffectiveCellRangeDimensions(ws, cr)
        self._results.addCellQueries(dims.cellsAccessed)
        self._results.addCellsWithData(dims.cellsPopulated)
        return CellRangeMetadata(
            dn,
            ws,
            cr,
            populated_height=dims.populated_height,
            populated_width=dims.populated_width,
            populated_min_col=dims.populated_min_col,
            populated_min_row=dims.populated_min_row,
        )

    def resolveRange(
        self,
        definedName: DefinedName | str | CellRangeMetadata,
    ) -> Optional[CellRangeMetadata]:
        """Resolve a defined name (or its string name) to its cell range,
        marking it as used. Returns None if the name is missing or damaged.

        An already-resolved CellRangeMetadata passes through unchanged (and is
        marked used), so callers can accept either form.
        """
        if isinstance(definedName, str):
            dn = self.getDefinedName(definedName)
            if dn is None:
                return None
            definedName = dn
        if isinstance(definedName, DefinedName):
            if (crm := self.peekRange(definedName)) is None:
                return None
            definedName = crm
        self.markUsed(definedName.definedName)
        return definedName

    def getSingleCell(
        self,
        definedName: DefinedName | str | CellRangeMetadata,
        *,
        row: int = -1,
        column: int = -1,
    ) -> Optional[CellType]:
        if (stuff := self.resolveRange(definedName)) is None:
            return None

        cr = stuff.cellRange
        ws = stuff.worksheet

        if not all(
            x is not None for x in (cr.min_row, cr.max_row, cr.min_col, cr.max_col)
        ):
            self._msg.error(
                f"Named range {stuff.definedName.name} has an invalid cell range {cr.bounds}.",
                MessageType.DevInfo,
                ref=excelDefinedNameRef(stuff.definedName),
            )
            return None

        shouldOverrideRow = row == -1 or stuff.maximum_height == 1
        shouldOverrideColumn = column == -1 or stuff.maximum_width == 1

        if shouldOverrideRow:
            row = cr.min_row
            if stuff.populated_height > 1:
                self._msg.warning(
                    f"Named range {stuff.definedName.name} has {stuff.populated_height} populated rows but no row was specified; using the first.",
                    MessageType.DevInfo,
                    ref=stuff,
                )

        if shouldOverrideColumn:
            column = cr.min_col
            if stuff.populated_width > 1:
                self._msg.warning(
                    f"Named range {stuff.definedName.name} has {stuff.populated_width} populated columns but no column was specified; using the first.",
                    MessageType.DevInfo,
                    ref=stuff,
                )

        if not (cr.min_row <= row <= cr.max_row):
            self._msg.warning(
                f"Row {row} has not been specified correctly.",
                MessageType.DevInfo,
                ref=stuff,
            )
            row = cr.min_row
        if not (cr.min_col <= column <= cr.max_col):
            self._msg.warning(
                f"Column {column} has not been specified correctly.",
                MessageType.DevInfo,
                ref=stuff,
            )
            column = cr.min_col

        cell = ws.cell(row=row, column=column)

        if cell is None or cell.value is None:
            return None

        if cell.value == EXCEL_PLACEHOLDER_VALUE:
            self._msg.error(
                f"Excel cell has an invalid stored value {EXCEL_PLACEHOLDER_VALUE}. Please check the Excel formula for this specific cell.",
                MessageType.ExcelParsing,
                ref=stuff.excelRef(cell),
            )
            return None
        return cell

    def value(
        self,
        definedName: DefinedName | str | CellRangeMetadata,
        *,
        row: int = -1,
        column: int = -1,
    ) -> CellValue:
        """The value of a (single-cell) named range, as a CellValue.

        A missing or damaged name yields a blank CellValue rather than an error.
        """
        return CellValue.fromCell(
            self.getSingleCell(definedName, row=row, column=column)
        )
