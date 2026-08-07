"""Cell ranges: named-range metadata plus geometry and iteration helpers.

CellRangeMetadata is the package's core value object — a defined name resolved
to a worksheet and cell range, with its populated extent precomputed.
XbrlConceptCellRangeMetadata additionally ties the range to a taxonomy concept.
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from typing import TYPE_CHECKING, NamedTuple

if TYPE_CHECKING:
    from typing import Self

    from mireport.taxonomy import Concept

from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.worksheet import Worksheet

from mireport.exceptions import OpenPyXlRelatedException
from mireport.xlsx_template_reader._constants import CellType
from mireport.xlsx_template_reader.util import excelCellRangeRef, excelCellRef


@dataclass(slots=True, eq=True, frozen=True)
class CellRangeMetadata:
    definedName: DefinedName
    worksheet: Worksheet
    cellRange: CellRange
    populated_width: int
    populated_height: int
    populated_min_col: int
    populated_min_row: int

    @property
    def maximum_width(self) -> int:
        return self.cellRange.max_col - self.cellRange.min_col + 1

    @property
    def maximum_height(self) -> int:
        return self.cellRange.max_row - self.cellRange.min_row + 1

    def contains(self, other: CellRangeMetadata) -> bool:
        """True if other is on the same worksheet and fully within this range."""
        return self.worksheet is other.worksheet and self.cellRange.issuperset(
            other.cellRange
        )

    def overlaps(self, other: CellRangeMetadata) -> bool:
        """True if other is on the same worksheet and shares any cells with this range."""
        return self.worksheet is other.worksheet and not self.cellRange.isdisjoint(
            other.cellRange
        )

    def excelRef(self, cell: CellType | None = None) -> str:
        """Excel reference for this range, or for a specific cell within it.

        e.g. 'Example sheet'!$A$5:$B$10 for the range, 'Example sheet'!$A$5
        when a cell is given.
        """
        if cell is not None:
            return excelCellRef(self.worksheet, cell)
        return excelCellRangeRef(self.worksheet, self.cellRange)

    def rows(self) -> Iterator[tuple[int, tuple[CellType, ...]]]:
        """Yield (row_number, cells) for each row of the range."""
        return iterRows(self.worksheet, self.cellRange)

    def cells(self) -> Iterator[CellType]:
        """Yield every cell of the range, row by row."""
        return iterCells(self.worksheet, self.cellRange)

    def cellsWithCoords(self) -> Iterator[tuple[int, int, CellType]]:
        """Yield (row_number, column_number, cell) for every cell of the range."""
        return iterCellsWithCoords(self.worksheet, self.cellRange)


@dataclass(slots=True, eq=True, frozen=True)
class XbrlConceptCellRangeMetadata(CellRangeMetadata):
    concept: Concept

    @classmethod
    def fromCellRangeMetadata(cls, holder: CellRangeMetadata, concept: Concept) -> Self:
        return cls(
            definedName=holder.definedName,
            worksheet=holder.worksheet,
            cellRange=holder.cellRange,
            populated_width=holder.populated_width,
            populated_height=holder.populated_height,
            populated_min_col=holder.populated_min_col,
            populated_min_row=holder.populated_min_row,
            concept=concept,
        )

    def conflictsWith(self, other: XbrlConceptCellRangeMetadata) -> bool:
        """True if the two ranges overlap in a way a hypercube table can't allow.

        Within a table, two concept ranges must be either disjoint or the exact
        same range shared by two reportable primary items. Any other overlap
        (partial, or a same range involving a dimension) is a conflict.
        """
        if not self.overlaps(other):
            return False
        shares_reportable_range = (
            self.concept.isReportable
            and other.concept.isReportable
            and self.cellRange.bounds == other.cellRange.bounds
        )
        return not shares_reportable_range


def iterRows(
    ws: Worksheet, cr: CellRange
) -> Iterator[tuple[int, tuple[CellType, ...]]]:
    """Yield (row_number, cells) for each row of the range."""
    if cr.min_row is None or cr.min_col is None:
        raise OpenPyXlRelatedException(
            f"Cell range bounds expected to be int but actually None {cr=}"
        )
    yield from enumerate(
        ws.iter_rows(
            min_row=cr.min_row,
            min_col=cr.min_col,
            max_row=cr.max_row,
            max_col=cr.max_col,
        ),
        start=cr.min_row,
    )


def iterCells(ws: Worksheet, cr: CellRange) -> Iterator[CellType]:
    """Yield every cell of the range, row by row."""
    for _, row in iterRows(ws, cr):
        yield from row


def iterCellsWithCoords(
    ws: Worksheet, cr: CellRange
) -> Iterator[tuple[int, int, CellType]]:
    """Yield (row_number, column_number, cell) for every cell of the range."""
    for rnum, row in iterRows(ws, cr):
        for cnum, cell in enumerate(row, start=cr.min_col):
            yield rnum, cnum, cell


class _CellRangeDimensions(NamedTuple):
    cellsAccessed: set[tuple[str, int, int]]
    cellsPopulated: set[tuple[str, int, int]]
    populated_width: int
    populated_height: int
    populated_min_col: int
    populated_min_row: int

    @property
    def countAccessed(self) -> int:
        return len(self.cellsAccessed)

    @property
    def countPopulated(self) -> int:
        return len(self.cellsPopulated)


def getEffectiveCellRangeDimensions(
    ws: Worksheet, cell_range: CellRange
) -> _CellRangeDimensions:
    cols_not_empty: set[int] = set()
    cols_with_none: set[int] = set()
    populated_rows: set[int] = set()
    populatedCellCount: set[tuple[str, int, int]] = set()
    cellCount: set[tuple[str, int, int]] = set()

    last_rnum = None
    empty_row = True
    sheetName = ws.title
    for rnum, cnum, cell in iterCellsWithCoords(ws, cell_range):
        cellCount.add((sheetName, rnum, cnum))
        if last_rnum is None:
            last_rnum = rnum

        if rnum != last_rnum:
            if not empty_row:
                populated_rows.add(last_rnum)
            last_rnum = rnum
            empty_row = True

        if cell.value is not None:
            populatedCellCount.add((sheetName, rnum, cnum))
            empty_row = False
            cols_not_empty.add(cnum)
        else:
            cols_with_none.add(cnum)
    if last_rnum is not None and not empty_row:
        populated_rows.add(last_rnum)

    definitely_empty_cols = cols_with_none - cols_not_empty
    total_cols = len(cols_not_empty.union(cols_with_none))
    populated_width = max(1, total_cols - len(definitely_empty_cols))
    populated_height = max(1, len(populated_rows))
    populated_min_col = min(cols_not_empty, default=None) or cell_range.min_col
    populated_min_row = min(populated_rows, default=None) or cell_range.min_row
    return _CellRangeDimensions(
        cellsAccessed=cellCount,
        cellsPopulated=populatedCellCount,
        populated_width=populated_width,
        populated_height=populated_height,
        populated_min_col=populated_min_col,
        populated_min_row=populated_min_row,
    )
