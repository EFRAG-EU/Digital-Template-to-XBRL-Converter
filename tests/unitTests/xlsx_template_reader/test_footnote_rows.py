"""Unit tests for _iterFootnoteRows row/boundary handling.

The footnote table layout: the text column delimits footnotes (a non-merged
cell starts a new footnote; merged continuation cells belong to the one
above), the ref column holds concept references for the current footnote, and
an optional dimension column qualifies the refs on its own row.
"""

from openpyxl import Workbook
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from mireport.conversionresults import ConversionResultsBuilder, Severity
from mireport.xlsx_template_reader._bindings import FootnoteBinding
from mireport.xlsx_template_reader._footnotes import (
    FootnoteFactCreator,
    _columnIndex,
    _iterFootnoteRows,
)
from mireport.xlsx_template_reader._messages import Messenger
from mireport.xlsx_template_reader._reader import WorkbookReader

_SHEET = "Sheet"


def _wb() -> Workbook:
    wb = Workbook()
    wb.active.title = _SHEET  # type: ignore[union-attr]
    return wb


def _crm(wb: Workbook, reader: WorkbookReader, name: str, ref: str):
    attr = f"{quote_sheetname(_SHEET)}!{absolute_coordinate(ref)}"
    wb.defined_names[name] = DefinedName(name, attr_text=attr)
    crm = reader.peekRange(wb.defined_names[name])
    assert crm is not None
    return crm


def _footnotes(
    wb: Workbook,
    table_ref: str,
    text_ref: str,
    ref_ref: str,
    dim_ref: str | None = None,
):
    """Return the (text, [(label, dim_text), ...]) pairs _iterFootnoteRows yields."""
    reader = WorkbookReader(wb, ConversionResultsBuilder(consoleOutput=False))
    table = _crm(wb, reader, "footnote_table", table_ref)
    text = _crm(wb, reader, "footnote_text", text_ref)
    ref = _crm(wb, reader, "footnote_ref_concept", ref_ref)
    origin = table.cellRange.min_col
    dim_col = None
    if dim_ref is not None:
        dim_col = _columnIndex(_crm(wb, reader, "footnote_ref_dim", dim_ref), origin)
    return [
        (text_value, [(label, dim) for label, dim, _ in label_cells])
        for text_value, label_cells in _iterFootnoteRows(
            table, _columnIndex(text, origin), _columnIndex(ref, origin), dim_col
        )
    ]


def test_refs_under_blank_text_do_not_leak_into_next_footnote():
    """Refs on rows whose text cell is blank belong to no footnote and must be
    discarded, not attributed to the next footnote."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    # Row 1: blank text cell with a stray ref; row 2: a real footnote.
    ws["B1"] = "LeakedRef"
    ws["A2"] = "Real footnote"
    ws["B2"] = "RealRef"

    footnotes = _footnotes(wb, "A1:B2", "A1:A2", "B1:B2")

    assert footnotes == [("Real footnote", [("RealRef", None)])]


def test_ref_in_otherwise_blank_table_yields_nothing():
    """A stray ref with no footnote text anywhere must also be discarded at the
    end of the table, not just when a later footnote starts."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["B1"] = "StrayRef"

    assert _footnotes(wb, "A1:B1", "A1:A1", "B1:B1") == []


def test_merged_text_block_collects_refs_from_all_its_rows():
    """Refs on every row of a merged text block belong to that footnote."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws.merge_cells("A1:A2")
    ws["A1"] = "Footnote one"
    ws["B1"] = "Ref1"
    ws["B2"] = "Ref2"
    ws["A3"] = "Footnote two"
    ws["B3"] = "Ref3"

    footnotes = _footnotes(wb, "A1:B3", "A1:A3", "B1:B3")

    assert footnotes == [
        ("Footnote one", [("Ref1", None), ("Ref2", None)]),
        ("Footnote two", [("Ref3", None)]),
    ]


def test_dimension_label_pairs_with_refs_on_its_own_row():
    """A dimension label qualifies the refs on its row only, not the whole block."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws.merge_cells("A1:A2")
    ws["A1"] = "Footnote"
    ws["B1"] = "Ref1"
    ws["C1"] = "Region"
    ws["B2"] = "Ref2"

    footnotes = _footnotes(wb, "A1:C2", "A1:A2", "B1:B2", dim_ref="C1:C2")

    assert footnotes == [("Footnote", [("Ref1", "Region"), ("Ref2", None)])]


def test_dimension_without_ref_on_its_row_is_ignored():
    """A dimension label on a row with no ref qualifies nothing — it does not
    attach to refs on other rows of the block, and yields no reference."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws.merge_cells("A1:A2")
    ws["A1"] = "Footnote"
    ws["B1"] = "Ref1"
    ws["C2"] = "Region"

    footnotes = _footnotes(wb, "A1:C2", "A1:A2", "B1:B2", dim_ref="C1:C2")

    assert footnotes == [("Footnote", [("Ref1", None)])]


def test_text_ref_and_dimension_values_are_stripped():
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "  Footnote  "
    ws["B1"] = "  Ref  "
    ws["C1"] = "  Region  "

    footnotes = _footnotes(wb, "A1:C1", "A1:A1", "B1:B1", dim_ref="C1:C1")

    assert footnotes == [("Footnote", [("Ref", "Region")])]


def test_whitespace_only_text_is_no_footnote():
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "   "
    ws["B1"] = "Ref"

    assert _footnotes(wb, "A1:B1", "A1:A1", "B1:B1") == []


def test_placeholder_text_is_no_footnote():
    """A '-' placeholder in the text column is a blank block, not a footnote."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "-"
    ws["B1"] = "Ref"

    assert _footnotes(wb, "A1:B1", "A1:A1", "B1:B1") == []


def test_placeholder_ref_is_skipped():
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Footnote"
    ws["B1"] = "-"

    assert _footnotes(wb, "A1:B1", "A1:A1", "B1:B1") == [("Footnote", [])]


def test_text_without_ref_yields_footnote_with_no_refs():
    """The iterator still yields a ref-less footnote; createFootnotes is the
    layer that warns about it and skips."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Lonely footnote"

    assert _footnotes(wb, "A1:B1", "A1:A1", "B1:B1") == [("Lonely footnote", [])]


def test_text_without_ref_warns_and_attaches_nothing():
    """createFootnotes warns about a footnote with no concept references and
    skips it without touching the report."""
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Lonely footnote"

    results = ConversionResultsBuilder(consoleOutput=False)
    reader = WorkbookReader(wb, results)
    binding = FootnoteBinding(
        table=_crm(wb, reader, "footnote_table", "A1:B1"),
        text=_crm(wb, reader, "footnote_text", "A1:A1"),
        ref=_crm(wb, reader, "footnote_ref_concept", "B1:B1"),
        ref_dimension=None,
    )
    creator = FootnoteFactCreator(
        report=None,  # type: ignore[arg-type] # skipped footnotes never touch the report
        msg=Messenger(results),
        binding=binding,
    )

    creator.createFootnotes()

    warnings = [
        str(m.messageText) for m in results.messages if m.severity is Severity.WARNING
    ]
    assert any("no concept references" in w for w in warnings), warnings


def test_placeholder_dimension_is_ignored():
    wb = _wb()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Footnote"
    ws["B1"] = "Ref"
    ws["C1"] = "-"

    footnotes = _footnotes(wb, "A1:C1", "A1:A1", "B1:B1", dim_ref="C1:C1")

    assert footnotes == [("Footnote", [("Ref", None)])]
