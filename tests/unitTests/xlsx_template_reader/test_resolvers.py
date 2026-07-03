"""Tests for the resolver layer: per-table XBRLTableResolver, one-shot resolver
functions, and the VSME taxonomy name aliases."""

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from mireport.conversionresults import ConversionResultsBuilder, Severity
from mireport.data.disclosures import VSME_DEFAULTS
from mireport.taxonomy import getTaxonomy
from mireport.xlsx_template_reader._binder import WorkbookBinder
from mireport.xlsx_template_reader._messages import Messenger
from mireport.xlsx_template_reader._ranges import XbrlConceptCellRangeMetadata
from mireport.xlsx_template_reader._reader import WorkbookReader
from mireport.xlsx_template_reader._resolvers import (
    ExcelCellBindingContext,
    XBRLTableResolver,
    resolveExternalValues,
    resolveFootnoteBinding,
    resolveNamedRangeTable,
)
from mireport.xlsx_template_reader.util import loadExcelFromPathOrFileLike

SAMPLE = (
    Path(__file__).parent.parent.parent
    / "data"
    / "VSME-Digital-Template-Sample-1.2.0.xlsx"
)

ALIASED_NAME = "IdentifierOfSitesInBiodiversitySensitiveAreasTypedAxis"
ALIAS_TARGET = "IdentifierOfSiteTypedAxis"


def _results() -> ConversionResultsBuilder:
    return ConversionResultsBuilder(consoleOutput=False)


@pytest.fixture(scope="module")
def reader():
    wb = loadExcelFromPathOrFileLike(SAMPLE)
    yield WorkbookReader(wb, _results())
    wb.close()


@pytest.fixture(scope="module")
def taxonomy(reader):
    return getTaxonomy(reader.value(VSME_DEFAULTS["entryPoint"]).as_str())


@pytest.fixture(scope="module")
def ctx(reader, taxonomy):
    return ExcelCellBindingContext(reader, Messenger(_results()), taxonomy)


@pytest.fixture(scope="module")
def bound(taxonomy):
    wb = loadExcelFromPathOrFileLike(SAMPLE)
    reader = WorkbookReader(wb, _results())
    try:
        yield WorkbookBinder(reader, taxonomy, _results()).bind()
    finally:
        wb.close()


@pytest.fixture(scope="module")
def rich_hypercube(taxonomy):
    """A hypercube with at least two reportable primary items and a dimension,
    plus those items — the raw material for classification scenarios."""
    for hc in sorted(taxonomy.hypercubes, key=str):
        pris = [c for c in taxonomy.getPrimaryItemsForHypercube(hc) if c.isReportable]
        dims = sorted(taxonomy.getDimensionsForHypercube(hc), key=str)
        if len(pris) >= 2 and dims:
            return hc, sorted(pris, key=str), dims
    pytest.skip("no hypercube with >=2 reportable primary items and a dimension")


def _classify(taxonomy, hypercube, placements, unit_map=None):
    """Build a synthetic worksheet with a table range (A1:F6) for the hypercube
    and one named range per (concept, cell_ref) placement, then resolve it.

    Returns (TableBinding | None, results).
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"

    def crm(name, ref, concept):
        attr = f"{quote_sheetname('S')}!{absolute_coordinate(ref)}"
        wb.defined_names[name] = DefinedName(name, attr_text=attr)
        resolved = reader.peekRange(wb.defined_names[name])
        assert resolved is not None
        return XbrlConceptCellRangeMetadata.fromCellRangeMetadata(resolved, concept)

    results = _results()
    reader = WorkbookReader(wb, results)
    table_crm = crm("tbl", "A1:F6", hypercube)
    candidates = [
        crm(f"r{i}", ref, concept) for i, (concept, ref) in enumerate(placements)
    ]
    resolver = XBRLTableResolver(
        ExcelCellBindingContext(reader, Messenger(results), taxonomy),
        unit_map=unit_map or {},
        candidates_by_ws={table_crm.worksheet: candidates},
        concepts_in_excel=frozenset(
            [hypercube, *(concept for concept, _ in placements)]
        ),
    )
    return resolver.resolve(table_crm), results


def _messages(results, severity):
    return [str(m.messageText) for m in results.messages if m.severity is severity]


class TestXBRLTableResolverClassification:
    def test_partitions_primary_items_dimensions_and_units(
        self, taxonomy, rich_hypercube
    ):
        hc, pris, dims = rich_hypercube
        pri, dim = pris[0], dims[0]
        binding, _ = _classify(taxonomy, hc, [(pri, "B2"), (dim, "C2")])
        assert binding is not None
        assert [c.concept for c in binding.primaryItems] == [pri]
        expected_typed = [dim] if dim.isTypedDimension else []
        expected_explicit = [dim] if dim.isExplicitDimension else []
        assert [c.concept for c in binding.typedDimensions] == expected_typed
        assert [c.concept for c in binding.explicitDimensions] == expected_explicit

    def test_unit_map_entry_for_primary_item_lands_in_units(
        self, taxonomy, rich_hypercube
    ):
        hc, pris, _dims = rich_hypercube
        pri = pris[0]
        # Any resolved range works as the unit holder; reuse a placement helper
        # by classifying once to get hold of a crm for the unit map.
        first, _ = _classify(taxonomy, hc, [(pri, "B2")])
        assert first is not None
        unit_holder = first.primaryItems[0]
        binding, _ = _classify(taxonomy, hc, [(pri, "B2")], unit_map={pri: unit_holder})
        assert binding is not None
        assert binding.units == [unit_holder]

    def test_partially_overlapping_ranges_reject_the_table(
        self, taxonomy, rich_hypercube
    ):
        hc, pris, _ = rich_hypercube
        binding, results = _classify(
            taxonomy, hc, [(pris[0], "B2:C2"), (pris[1], "C2:D2")]
        )
        assert binding is None
        assert any(
            "neither the same nor disjoint" in e
            for e in _messages(results, Severity.ERROR)
        )

    def test_identical_ranges_for_two_reportables_are_allowed(
        self, taxonomy, rich_hypercube
    ):
        hc, pris, _ = rich_hypercube
        binding, _ = _classify(taxonomy, hc, [(pris[0], "B2:C2"), (pris[1], "B2:C2")])
        assert binding is not None
        assert {c.concept for c in binding.primaryItems} == {pris[0], pris[1]}

    def test_unpermitted_concept_inside_table_is_extra(self, taxonomy, rich_hypercube):
        hc, pris, _dims = rich_hypercube
        permitted = taxonomy.getDimensionsForHypercube(hc).union(
            taxonomy.getPrimaryItemsForHypercube(hc)
        )
        outsider = next(
            c
            for c in sorted(taxonomy.concepts, key=str)
            if c.isReportable and c not in permitted
        )
        binding, results = _classify(taxonomy, hc, [(pris[0], "B2"), (outsider, "C3")])
        assert binding is not None
        assert outsider not in {c.concept for c in binding.primaryItems}
        assert any(
            "Extra named ranges" in w for w in _messages(results, Severity.WARNING)
        )

    def test_missing_permitted_concepts_warn(self, taxonomy, rich_hypercube):
        hc, pris, _ = rich_hypercube
        # Only one of the >=2 primary items is placed, so the rest are missing.
        _, results = _classify(taxonomy, hc, [(pris[0], "B2")])
        assert any(
            "have not been found" in w for w in _messages(results, Severity.WARNING)
        )


def _footnote_ctx(taxonomy, ranges):
    """Build a tiny in-memory workbook holding the given footnote named ranges
    (name -> Excel coords), plus its resolver context."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    for name, coords in ranges.items():
        attr = f"{quote_sheetname(ws.title)}!{absolute_coordinate(coords)}"
        wb.defined_names.add(DefinedName(name, attr_text=attr))
    results = _results()
    reader = WorkbookReader(wb, results)
    return ExcelCellBindingContext(reader, Messenger(results), taxonomy), results


class TestResolveFootnoteBindingColumnWidths:
    """The footnote reader only supports single-column text/ref/dimension
    ranges; wider ranges must warn up front and fall back to the first column."""

    def test_single_column_ranges_bind_without_warning(self, taxonomy):
        ctx, results = _footnote_ctx(
            taxonomy,
            {
                "footnote_table": "B1:D2",
                "footnote_text": "B1:B2",
                "footnote_ref_dimension": "C1:C2",
                "footnote_ref_concept": "D1:D2",
            },
        )
        assert resolveFootnoteBinding(ctx) is not None
        assert not _messages(results, Severity.WARNING)

    def test_wide_ref_range_warns_and_still_binds(self, taxonomy):
        ctx, results = _footnote_ctx(
            taxonomy,
            {
                "footnote_table": "B1:D2",
                "footnote_text": "B1:B2",
                "footnote_ref_concept": "C1:D2",
            },
        )
        assert resolveFootnoteBinding(ctx) is not None
        warnings = _messages(results, Severity.WARNING)
        assert any(
            "footnote_ref_concept" in w and "first column" in w for w in warnings
        ), warnings

    def test_wide_text_and_dimension_ranges_each_warn(self, taxonomy):
        ctx, results = _footnote_ctx(
            taxonomy,
            {
                "footnote_table": "B1:F2",
                "footnote_text": "B1:C2",
                "footnote_ref_dimension": "D1:E2",
                "footnote_ref_concept": "F1:F2",
            },
        )
        assert resolveFootnoteBinding(ctx) is not None
        warnings = _messages(results, Severity.WARNING)
        assert any("footnote_text" in w for w in warnings), warnings
        assert any("footnote_ref_dimension" in w for w in warnings), warnings


class TestOneShotResolvers:
    def test_footnotes_none_when_unconfigured(self, ctx):
        # The 1.2.0 sample has no footnote named ranges: silently nothing.
        assert resolveFootnoteBinding(ctx) is None

    def test_named_range_table_none_when_unconfigured(self, ctx):
        assert (
            resolveNamedRangeTable(
                ctx,
                label="Bogus",
                container_name="bogus_container",
                required_sub_names=("bogus_a", "bogus_b"),
                context="Testing.",
            )
            is None
        )


def _external_values_ctx(cells, taxonomy):
    """Build a tiny in-memory workbook whose template_external_values range
    holds the given cell values (one per row), plus its resolver context."""
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    for i, v in enumerate(cells, start=1):
        ws.cell(row=i, column=1).value = v
    wb.defined_names.add(
        DefinedName(
            "template_external_values",
            attr_text=f"'{ws.title}'!$A$1:$A${max(1, len(cells))}",
        )
    )
    results = _results()
    reader = WorkbookReader(wb, results)
    return ExcelCellBindingContext(reader, Messenger(results), taxonomy), results


def _warnings(results):
    from mireport.conversionresults import Severity

    return [m for m in results.messages if m.severity is Severity.WARNING]


class TestResolveExternalValues:
    @pytest.fixture(scope="class")
    def textblock(self, taxonomy):
        return next(c for c in sorted(taxonomy.concepts, key=str) if c.isTextblock)

    def test_known_concept_name_resolved(self, taxonomy, textblock):
        ctx, results = _external_values_ctx([textblock.qname.localName], taxonomy)
        assert resolveExternalValues(ctx) == frozenset({textblock})
        assert not _warnings(results)

    def test_blank_and_placeholder_cells_skipped_silently(self, taxonomy):
        ctx, results = _external_values_ctx([None, "-", "   ", "#VALUE!"], taxonomy)
        assert resolveExternalValues(ctx) == frozenset()
        assert not _warnings(results)

    def test_non_string_cells_warn_and_are_excluded(self, taxonomy):
        # Numbers/booleans can't name a concept: they now stringify, fail the
        # lookup and warn (previously they were skipped silently).
        ctx, results = _external_values_ctx([42, 3.14, True], taxonomy)
        assert resolveExternalValues(ctx) == frozenset()
        assert len(_warnings(results)) == 3

    def test_unknown_name_warns_and_is_excluded(self, taxonomy):
        ctx, results = _external_values_ctx(["NoSuchConceptXYZ"], taxonomy)
        assert resolveExternalValues(ctx) == frozenset()
        assert len(_warnings(results)) == 1

    def test_ambiguous_name_warns_and_is_excluded(self):
        from mireport.exceptions import AmbiguousComponentException

        class AmbiguousTaxonomy:
            def resolveConcept(self, text, **kwargs):
                raise AmbiguousComponentException(f"'{text}' is ambiguous")

        ctx, results = _external_values_ctx(["SharedLabel"], AmbiguousTaxonomy())
        assert resolveExternalValues(ctx) == frozenset()
        assert len(_warnings(results)) == 1


class TestBindOrderIsDeterministic:
    def test_concept_map_keys_are_name_sorted(self, bound):
        """DefinedName hashes by identity, so iterating the reader's unused-name
        set directly gives a different order every run — which leaks into fact
        and message ordering. bind() must impose a stable (name-sorted) order."""
        names = [dn.name for dn in bound.concept_map]
        assert names == sorted(names)


class TestTaxonomyNameAliases:
    def test_bind_applies_alias(self, bound):
        """The aliased defined name must bind to the alias-target concept,
        whether it ended up in the concept_map or inside a table binding."""
        all_ranges = list(bound.concept_map.values())
        for table_binding in bound.tables:
            all_ranges.extend(table_binding.conceptRanges)
        aliased = [crm for crm in all_ranges if crm.definedName.name == ALIASED_NAME]
        assert aliased, f"sample should contain the {ALIASED_NAME} named range"
        for crm in aliased:
            assert crm.concept.qname.localName == ALIAS_TARGET
