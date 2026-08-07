"""Table EE cells must use the same label-resolution chain as simple facts:
exact standard label -> configured cell-value alias -> closest EE-domain match.

Historically the table path used a bare label lookup (the since-removed
getConceptForLabel), so table cells with slightly-off labels errored where
simple-fact cells resolved.
"""

from collections import defaultdict

import pytest
from openpyxl import Workbook
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from mireport.conversionresults import ConversionResultsBuilder, Severity
from mireport.data.disclosures import VSME_DEFAULTS
from mireport.exceptions import AmbiguousComponentException
from mireport.report import InlineReport
from mireport.taxonomy import getTaxonomy, listTaxonomies
from mireport.xlsx_template_reader._bindings import TableBinding, WorkbookBindings
from mireport.xlsx_template_reader._config import ConverterConfig
from mireport.xlsx_template_reader._enumerations import resolveMemberByLabel
from mireport.xlsx_template_reader._messages import Messenger
from mireport.xlsx_template_reader._ranges import XbrlConceptCellRangeMetadata
from mireport.xlsx_template_reader._reader import WorkbookReader
from mireport.xlsx_template_reader._tables import TableFactCreator
from mireport.xlsx_template_reader._units import UnitResolver

_SHEET = "S"


@pytest.fixture(scope="module")
def taxonomy():
    entry_point = next(ep for ep in listTaxonomies() if "vsme" in ep.lower())
    return getTaxonomy(entry_point)


@pytest.fixture(scope="module")
def ee_concept(taxonomy):
    return next(
        c
        for c in sorted(taxonomy.concepts, key=str)
        if c.isEnumerationSingle and c.isReportable and c.getEEDomain()
    )


def _table_env(taxonomy, ee_concept, cell_value):
    """A one-cell 'table' whose only primary item is the EE concept."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _SHEET
    ws["A1"] = cell_value
    attr = f"{quote_sheetname(_SHEET)}!{absolute_coordinate('A1')}"
    wb.defined_names["pri"] = DefinedName("pri", attr_text=attr)

    results = ConversionResultsBuilder(consoleOutput=False)
    reader = WorkbookReader(wb, results)
    crm = reader.peekRange(wb.defined_names["pri"])
    assert crm is not None
    pri = XbrlConceptCellRangeMetadata.fromCellRangeMetadata(crm, ee_concept)
    table_binding = TableBinding(
        table=pri,
        primaryItems=[pri],
        explicitDimensions=[],
        typedDimensions=[],
        units=[],
    )
    bindings = WorkbookBindings(
        concept_map={},
        tables=[table_binding],
        unit_map={},
        preset_dims=defaultdict(dict),
        has_external_value=frozenset(),
        footnote=None,
    )
    report = InlineReport(taxonomy, None)
    config = ConverterConfig.fromDefaults(VSME_DEFAULTS, taxonomy)
    msg = Messenger(results)
    units = UnitResolver(report, config, msg, reader, {})
    creator = TableFactCreator(report, reader, msg, config, units, bindings)
    return creator, results


def _messages(results, severity):
    return [str(m.messageText) for m in results.messages if m.severity is severity]


def test_exact_label_resolves_without_ee_errors(taxonomy, ee_concept):
    member = ee_concept.getEEDomain()[0]
    creator, results = _table_env(taxonomy, ee_concept, member.getStandardLabel())
    creator.createTableFacts()
    errors = _messages(results, Severity.ERROR)
    assert not any("Unable to find EE concept" in e for e in errors), errors


def test_fuzzy_label_resolves_with_closest_match_warning(taxonomy, ee_concept):
    """A typo'd member label must fuzzy-resolve (with the standard 'closest
    match' warning) instead of erroring, as it already does for simple facts."""
    member = ee_concept.getEEDomain()[0]
    typo = member.getStandardLabel() + " x"
    creator, results = _table_env(taxonomy, ee_concept, typo)
    creator.createTableFacts()

    errors = _messages(results, Severity.ERROR)
    assert not any("Unable to find EE concept" in e for e in errors), errors
    warnings = _messages(results, Severity.WARNING)
    assert any("closest match" in w for w in warnings), warnings


# ---------------------------------------------------------------------------
# Explicit-dimension member resolution (scoped to the dimension's domain)
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def config(taxonomy):
    return ConverterConfig.fromDefaults(VSME_DEFAULTS, taxonomy)


@pytest.fixture(scope="module")
def dim_concept(taxonomy):
    dim = next(
        (
            c
            for c in sorted(taxonomy.concepts, key=str)
            if c.isExplicitDimension
            and taxonomy.getDomainMembersForExplicitDimension(c)
        ),
        None,
    )
    if dim is None:
        pytest.skip("vsme taxonomy has no explicit dimension with domain members")
    return dim


@pytest.fixture(scope="module")
def pri_concept(taxonomy):
    """A plain (non-numeric, non-EE, non-date) reportable concept so the row
    processing exercises only the explicit-dimension path."""
    return next(
        c
        for c in sorted(taxonomy.concepts, key=str)
        if c.isReportable
        and not c.isNumeric
        and not c.isEnumerationSingle
        and not c.isEnumerationSet
        and not c.isDate
    )


def _dim_table_env(taxonomy, pri_concept, dim_concept, pri_value, dim_value):
    """A one-row table: primary item at A1, explicit-dimension cell at B1."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = _SHEET
    ws["A1"] = pri_value
    ws["B1"] = dim_value
    for name, coord in (("tbl", "A1:B1"), ("pri", "A1"), ("dim", "B1")):
        attr = f"{quote_sheetname(_SHEET)}!{absolute_coordinate(coord)}"
        wb.defined_names[name] = DefinedName(name, attr_text=attr)

    results = ConversionResultsBuilder(consoleOutput=False)
    reader = WorkbookReader(wb, results)

    def crm_for(name, concept):
        crm = reader.peekRange(wb.defined_names[name])
        assert crm is not None
        return XbrlConceptCellRangeMetadata.fromCellRangeMetadata(crm, concept)

    table_binding = TableBinding(
        table=crm_for("tbl", pri_concept),
        primaryItems=[crm_for("pri", pri_concept)],
        explicitDimensions=[crm_for("dim", dim_concept)],
        typedDimensions=[],
        units=[],
    )
    bindings = WorkbookBindings(
        concept_map={},
        tables=[table_binding],
        unit_map={},
        preset_dims=defaultdict(dict),
        has_external_value=frozenset(),
        footnote=None,
    )
    report = InlineReport(taxonomy, None)
    config = ConverterConfig.fromDefaults(VSME_DEFAULTS, taxonomy)
    msg = Messenger(results)
    units = UnitResolver(report, config, msg, reader, {})
    creator = TableFactCreator(report, reader, msg, config, units, bindings)
    return creator, results


class TestExplicitDimensionResolution:
    def test_valid_member_label_sets_dimension(
        self, taxonomy, config, dim_concept, pri_concept
    ):
        domain = taxonomy.getDomainMembersForExplicitDimension(dim_concept)
        label = None
        for member in sorted(domain):
            candidate = member.getStandardLabel()
            if candidate is None:
                continue
            try:
                match = resolveMemberByLabel(
                    taxonomy, config, candidate, dimension=dim_concept
                )
            except AmbiguousComponentException:
                continue
            if match is not None and match.concept == member:
                label = candidate
                break
        if label is None:
            pytest.skip("no resolvable domain-member label for this dimension")

        creator, results = _dim_table_env(
            taxonomy, pri_concept, dim_concept, "some text", label
        )
        creator.createTableFacts()
        errors = _messages(results, Severity.ERROR)
        assert not any("Required explicit dimension" in e for e in errors), errors

    def test_ambiguous_dimension_value_errors_instead_of_crashing(
        self, taxonomy, dim_concept, pri_concept, monkeypatch
    ):
        """AmbiguousComponentException from the label chain must become an
        error message naming the candidates, not abort the conversion."""
        import mireport.xlsx_template_reader._tables as tables_mod

        class _Candidate:
            def __init__(self, qname):
                self.qname = qname

        def raiser(*args, **kwargs):
            raise AmbiguousComponentException(
                "ambiguous",
                candidates=(_Candidate("vsme:One"), _Candidate("vsme:Two")),
            )

        monkeypatch.setattr(tables_mod, "resolveMemberByLabel", raiser)
        creator, results = _dim_table_env(
            taxonomy, pri_concept, dim_concept, "some text", "Shared label"
        )
        creator.createTableFacts()  # must not raise
        errors = _messages(results, Severity.ERROR)
        assert any("vsme:One" in e and "vsme:Two" in e for e in errors), errors

    def test_out_of_domain_label_does_not_set_dimension(
        self, taxonomy, config, dim_concept, pri_concept
    ):
        """An exact label match outside the dimension's domain must no longer
        be accepted as the dimension value."""
        domain = taxonomy.getDomainMembersForExplicitDimension(dim_concept)
        outsider_label = None
        for concept in sorted(taxonomy.concepts, key=str):
            if concept in domain:
                continue
            for candidate in concept.getAllStandardLabels():
                try:
                    unscoped = resolveMemberByLabel(taxonomy, config, candidate)
                except AmbiguousComponentException:
                    continue
                if unscoped is not None and unscoped.concept == concept:
                    outsider_label = candidate
                    break
            if outsider_label is not None:
                break
        if outsider_label is None:
            pytest.skip("no out-of-domain label resolves unscoped")

        creator, results = _dim_table_env(
            taxonomy, pri_concept, dim_concept, "some text", outsider_label
        )
        creator.createTableFacts()
        errors = _messages(results, Severity.ERROR)
        assert any("Required explicit dimension" in e for e in errors), errors
