"""Tests for WorkbookReader.getSingleCell."""

from openpyxl import Workbook
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from mireport.conversionresults import ConversionResultsBuilder, Severity
from mireport.xlsx_template_reader._reader import (
    EXCEL_PLACEHOLDER_VALUE,
    WorkbookReader,
)

_SHEET = "Sheet"


def _results() -> ConversionResultsBuilder:
    return ConversionResultsBuilder(consoleOutput=False)


def _wb() -> Workbook:
    wb = Workbook()
    wb.active.title = _SHEET  # type: ignore[union-attr]
    return wb


def _ws(wb: Workbook):
    assert wb.active is not None
    return wb.active


def _add_range(wb: Workbook, name: str, ref: str) -> None:
    """Add a named range; ref is like 'A1' or 'A1:C3' (sheet is always _SHEET)."""
    attr = f"{quote_sheetname(_SHEET)}!{absolute_coordinate(ref)}"
    wb.defined_names[name] = DefinedName(name, attr_text=attr)


def _reader(
    wb: Workbook, results: ConversionResultsBuilder | None = None
) -> WorkbookReader:
    return WorkbookReader(wb, results if results is not None else _results())


def _dev_messages(results: ConversionResultsBuilder) -> str:
    msgs = "\n".join(str(m) for m in results.developerMessages)
    return f"dev messages ({len(results.developerMessages)}):\n{msgs}"


def _has_dev_warning(results: ConversionResultsBuilder) -> bool:
    return any(m.severity == Severity.WARNING for m in results.developerMessages)


class TestGetSingleCellBasic:
    def test_returns_cell_with_value(self):
        wb = _wb()
        _ws(wb)["A1"] = "hello"
        _add_range(wb, "r", "A1")
        cell = _reader(wb).getSingleCell("r")
        assert cell is not None
        assert cell.value == "hello"

    def test_returns_none_for_empty_cell(self):
        wb = _wb()
        _add_range(wb, "r", "A1")
        assert _reader(wb).getSingleCell("r") is None

    def test_returns_none_for_unknown_name(self):
        wb = _wb()
        assert _reader(wb).getSingleCell("does_not_exist") is None

    def test_returns_none_for_placeholder_value(self):
        wb = _wb()
        _ws(wb)["A1"] = EXCEL_PLACEHOLDER_VALUE
        _add_range(wb, "r", "A1")
        assert _reader(wb).getSingleCell("r") is None


class TestGetSingleCellMultiRow:
    def test_uses_effective_min_row_when_one_row_populated(self):
        """Range A1:A3 but only A2 has data — should return A1."""
        wb = _wb()
        _ws(wb)["A2"] = "value"
        _add_range(wb, "r", "A1:A3")
        cell = _reader(wb).getSingleCell("r")
        assert cell is None

    def test_warns_and_falls_back_to_min_row_when_multiple_rows_populated(self):
        wb = _wb()
        _ws(wb)["A1"] = "first"
        _ws(wb)["A3"] = "third"
        _add_range(wb, "r", "A1:A3")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r")
        assert cell is not None, _dev_messages(results)
        assert cell.row == 1, _dev_messages(results)
        assert _has_dev_warning(results), _dev_messages(results)


class TestGetSingleCellMultiCol:
    def test_uses_effective_min_col_when_one_col_populated(self):
        """Range A1:C1 but only B1 has data — should return A1."""
        wb = _wb()
        _ws(wb)["B1"] = "value"
        _add_range(wb, "r", "A1:C1")
        cell = _reader(wb).getSingleCell("r")
        assert cell is None

    def test_warns_and_falls_back_to_min_col_when_multiple_cols_populated(self):
        wb = _wb()
        _ws(wb)["A1"] = "first"
        _ws(wb)["C1"] = "third"
        _add_range(wb, "r", "A1:C1")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r")
        assert cell is not None, _dev_messages(results)
        assert cell.column == 1, _dev_messages(results)
        assert _has_dev_warning(results), _dev_messages(results)


class TestGetSingleCellRowColParams:
    def test_row_param_selects_specific_row(self):
        wb = _wb()
        _ws(wb)["A2"] = "row2"
        _ws(wb)["A3"] = "row3"
        _add_range(wb, "r", "A1:A3")
        cell = _reader(wb).getSingleCell("r", row=3)
        assert cell is not None
        assert cell.value == "row3"

    def test_column_param_selects_specific_column(self):
        wb = _wb()
        _ws(wb)["B1"] = "colB"
        _ws(wb)["C1"] = "colC"
        _add_range(wb, "r", "A1:C1")
        cell = _reader(wb).getSingleCell("r", column=3)
        assert cell is not None
        assert cell.value == "colC"

    def test_out_of_bounds_row_logs_warning(self):
        wb = _wb()
        _ws(wb)["A1"] = "value"
        _add_range(wb, "r", "A1:A2")
        results = _results()
        _reader(wb, results).getSingleCell("r", row=99)
        assert _has_dev_warning(results), _dev_messages(results)

    def test_out_of_bounds_column_logs_warning(self):
        wb = _wb()
        _ws(wb)["A1"] = "value"
        _add_range(wb, "r", "A1:B1")
        results = _results()
        _reader(wb, results).getSingleCell("r", column=99)
        assert _has_dev_warning(results), _dev_messages(results)


class TestGetSingleCellMergedCells:
    def test_returns_anchor_cell_value_when_named_range_is_anchor(self):
        """Named range is the anchor cell of a horizontal merge — value is readable."""
        wb = _wb()
        _ws(wb).merge_cells("A1:C1")
        _ws(wb)["A1"] = "merged"
        _add_range(wb, "r", "A1")
        cell = _reader(wb).getSingleCell("r")
        assert cell is not None
        assert cell.value == "merged"

    def test_returns_anchor_when_named_range_spans_merged_region(self):
        """Named range covers the full merged region A1:C1; populated_width collapses to 1."""
        wb = _wb()
        _ws(wb).merge_cells("A1:C1")
        _ws(wb)["A1"] = "merged"
        _add_range(wb, "r", "A1:C1")
        cell = _reader(wb).getSingleCell("r")
        assert cell is not None
        assert cell.value == "merged"

    def test_returns_none_for_slave_merged_cell(self):
        """Named range points at B1, a slave cell in merge A1:C1 — value is None."""
        wb = _wb()
        _ws(wb).merge_cells("A1:C1")
        _ws(wb)["A1"] = "merged"
        _add_range(wb, "r", "B1")
        assert _reader(wb).getSingleCell("r") is None

    def test_returns_anchor_when_named_range_spans_vertical_merge(self):
        """Named range covers merged A1:A3; populated_height collapses to 1."""
        wb = _wb()
        _ws(wb).merge_cells("A1:A3")
        _ws(wb)["A1"] = "vertical"
        _add_range(wb, "r", "A1:A3")
        cell = _reader(wb).getSingleCell("r")
        assert cell is not None
        assert cell.value == "vertical"

    def test_out_of_bounds_column_warns_and_falls_back_to_anchor(self):
        """column beyond merged A1:C1 range logs a warning and falls back to A1."""
        wb = _wb()
        _ws(wb).merge_cells("A1:C1")
        _ws(wb)["A1"] = "merged"
        _add_range(wb, "r", "A1:C1")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r", column=99)
        assert _has_dev_warning(results), _dev_messages(results)
        assert cell is not None
        assert cell.value == "merged"

    def test_horizontal_merges_per_row_collapse_width_warn_on_rows(self):
        """A1:C3 with each row merged horizontally: populated_width=1 but 3 rows have data.
        No column warning (only one effective column); row warning fires because no row given."""
        wb = _wb()
        ws = _ws(wb)
        ws.merge_cells("A1:C1")
        ws.merge_cells("A2:C2")
        ws.merge_cells("A3:C3")
        ws["A1"] = "r1"
        ws["A2"] = "r2"
        ws["A3"] = "r3"
        _add_range(wb, "r", "A1:C3")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r")
        assert _has_dev_warning(results), _dev_messages(results)
        assert cell is not None
        assert cell.value == "r1"

    def test_no_warning_when_row_specified_and_columns_collapsed_by_merge(self):
        """A1:C3 with each row merged horizontally: populated_width=1 so no column warning even
        though no column is given. Row is specified and in-bounds so no row warning either."""
        wb = _wb()
        ws = _ws(wb)
        ws.merge_cells("A1:C1")
        ws.merge_cells("A2:C2")
        ws.merge_cells("A3:C3")
        ws["A1"] = "r1"
        ws["A2"] = "r2"
        ws["A3"] = "r3"
        _add_range(wb, "r", "A1:C3")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r", row=2)
        assert not _has_dev_warning(results), _dev_messages(results)
        assert cell is not None
        assert cell.value == "r2"

    def test_no_warning_when_column_specified_and_rows_collapsed_by_merge(self):
        """A1:C3 with each column merged vertically: populated_height=1 so no row warning even
        though no row is given. Column is specified and in-bounds so no column warning either."""
        wb = _wb()
        ws = _ws(wb)
        ws.merge_cells("A1:A3")
        ws.merge_cells("B1:B3")
        ws.merge_cells("C1:C3")
        ws["A1"] = "cA"
        ws["B1"] = "cB"
        ws["C1"] = "cC"
        _add_range(wb, "r", "A1:C3")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r", column=2)
        assert not _has_dev_warning(results), _dev_messages(results)
        assert cell is not None
        assert cell.value == "cB"

    def test_vertical_merges_per_col_collapse_height_warn_on_cols(self):
        """A1:C3 with each column merged vertically: populated_height=1 but 3 columns have data.
        No row warning (only one effective row); column warning fires because no column given."""
        wb = _wb()
        ws = _ws(wb)
        ws.merge_cells("A1:A3")
        ws.merge_cells("B1:B3")
        ws.merge_cells("C1:C3")
        ws["A1"] = "cA"
        ws["B1"] = "cB"
        ws["C1"] = "cC"
        _add_range(wb, "r", "A1:C3")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r")
        assert _has_dev_warning(results), _dev_messages(results)
        assert cell is not None
        assert cell.value == "cA"

    def test_out_of_bounds_row_warns_and_falls_back_to_anchor(self):
        """row beyond merged A1:A3 range logs a warning and falls back to A1."""
        wb = _wb()
        _ws(wb).merge_cells("A1:A3")
        _ws(wb)["A1"] = "vertical"
        _add_range(wb, "r", "A1:A3")
        results = _results()
        cell = _reader(wb, results).getSingleCell("r", row=99)
        assert _has_dev_warning(results), _dev_messages(results)
        assert cell is not None
        assert cell.value == "vertical"
