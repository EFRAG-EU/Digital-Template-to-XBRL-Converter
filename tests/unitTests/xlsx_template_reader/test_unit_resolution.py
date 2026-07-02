"""Tests for the middle of UnitResolver.setUnitForName's resolution chain:

    unit named range (direct/parenthesised/config-corrected text)
    -> per-concept config override -> taxonomy required units
    -> config complex units -> data-type/UTR fallbacks

The endpoints (getSimpleUnit, setFallbackUnitForName) are covered in
test_fact_creator_characterization.py; these tests pin the branching between
them, which nothing else exercises directly.
"""

import pytest
from openpyxl import Workbook
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from mireport.conversionresults import ConversionResultsBuilder, Severity
from mireport.data.disclosures import VSME_DEFAULTS
from mireport.report import InlineReport
from mireport.taxonomy import getTaxonomy, listTaxonomies
from mireport.xlsx_template_reader._config import ConverterConfig
from mireport.xlsx_template_reader._messages import Messenger
from mireport.xlsx_template_reader._ranges import XbrlConceptCellRangeMetadata
from mireport.xlsx_template_reader._reader import WorkbookReader
from mireport.xlsx_template_reader._units import UnitResolver

_SHEET = "S"


@pytest.fixture(scope="module")
def taxonomy():
    entry_point = next(ep for ep in listTaxonomies() if "vsme" in ep.lower())
    return getTaxonomy(entry_point)


@pytest.fixture(scope="module")
def config(taxonomy):
    return ConverterConfig.fromDefaults(VSME_DEFAULTS, taxonomy)


@pytest.fixture(scope="module")
def energy_concept(taxonomy):
    """A reportable numeric concept whose data type accepts MWh but not kg."""
    mwh = taxonomy.QNameMaker.fromString("utr:MWh")
    kg = taxonomy.QNameMaker.fromString("utr:kg")
    concept = next(
        (
            c
            for c in sorted(taxonomy.concepts, key=str)
            if c.isReportable
            and c.isNumeric
            and taxonomy.UTR.valid(c.dataType, mwh)
            and not taxonomy.UTR.valid(c.dataType, kg)
        ),
        None,
    )
    if concept is None:
        pytest.skip("taxonomy has no MWh-but-not-kg concept")
    return concept


class _Env:
    """A concept cell (A1) and unit cell (B1) wired through a real reader."""

    def __init__(self, taxonomy, config, concept, unit_text, *, with_unit_map=True):
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = _SHEET
        ws["A1"] = 123
        ws["B1"] = unit_text
        for name, ref in (("c", "A1"), ("c_unit", "B1")):
            attr = f"{quote_sheetname(_SHEET)}!{absolute_coordinate(ref)}"
            wb.defined_names[name] = DefinedName(name, attr_text=attr)

        self.results = ConversionResultsBuilder(consoleOutput=False)
        reader = WorkbookReader(wb, self.results)

        def crm(name):
            resolved = reader.peekRange(wb.defined_names[name])
            assert resolved is not None
            return XbrlConceptCellRangeMetadata.fromCellRangeMetadata(resolved, concept)

        self.holder = crm("c")
        self.unit_holder = crm("c_unit")
        self.report = InlineReport(taxonomy, None)
        unit_map = {concept: self.unit_holder} if with_unit_map else {}
        self.resolver = UnitResolver(
            self.report, config, Messenger(self.results), reader, unit_map
        )
        self.fb = self.report.getFactBuilder().setConcept(concept)

    def warnings(self):
        return [
            str(m.messageText)
            for m in self.results.messages
            if m.severity is Severity.WARNING
        ]


class TestUnitFromNamedRange:
    def test_valid_unit_for_data_type_is_used(self, taxonomy, config, energy_concept):
        env = _Env(taxonomy, config, energy_concept, "MWh")
        assert env.resolver.setUnitForName(env.holder, env.fb) is True
        assert env.fb._aspects["units"].localName == "MWh"

    def test_wrong_unit_via_specified_holder_warns_and_fails(
        self, taxonomy, config, energy_concept
    ):
        """Table path: a shared-with-nobody unit range with a wrong-type unit
        means the fact cannot be created."""
        env = _Env(taxonomy, config, energy_concept, "kg")
        ok = env.resolver.setUnitForName(
            env.holder,
            env.fb,
            specifiedUnitHolder=env.unit_holder,
            sharedRange=False,
        )
        assert ok is False
        assert any("not matching data type" in w for w in env.warnings())

    def test_wrong_unit_via_shared_range_fails_silently(
        self, taxonomy, config, energy_concept
    ):
        """A unit range shared between primary items legitimately holds units
        that fit only some of them: no warning for the misfits."""
        env = _Env(taxonomy, config, energy_concept, "kg")
        ok = env.resolver.setUnitForName(
            env.holder,
            env.fb,
            specifiedUnitHolder=env.unit_holder,
            sharedRange=True,
        )
        assert ok is False
        assert not any("not matching data type" in w for w in env.warnings())

    def test_wrong_unit_via_unit_map_falls_back(self, taxonomy, config, energy_concept):
        """Simple-fact path: a wrong-type unit in the unit range falls through
        to the data-type/UTR fallback instead of failing."""
        env = _Env(taxonomy, config, energy_concept, "kg")
        assert env.resolver.setUnitForName(env.holder, env.fb) is True
        assert "units" in env.fb._aspects
        assert env.fb._aspects["units"].localName != "kg"

    def test_empty_unit_cell_fails(self, taxonomy, config, energy_concept):
        env = _Env(taxonomy, config, energy_concept, None)
        assert env.resolver.setUnitForName(env.holder, env.fb) is False


class TestConfiguredConceptUnit:
    def test_unresolvable_text_uses_concepts_to_units_override(self, taxonomy, config):
        concept, unit = next(iter(config.conceptsToUnits.items()))
        assert taxonomy.UTR.valid(concept.dataType, unit), (
            "config override must be valid for this test to be meaningful"
        )
        env = _Env(taxonomy, config, concept, "wibbles per parsec")
        assert env.resolver.setUnitForName(env.holder, env.fb) is True
        assert env.fb._aspects["units"] == unit


class TestNoUnitRange:
    def test_single_required_unit_is_used(self, taxonomy, config):
        concept = next(
            (
                c
                for c in sorted(taxonomy.concepts, key=str)
                if c.isReportable
                and c.isNumeric
                and (units := c.getRequiredUnitQNames()) is not None
                and len(units) == 1
            ),
            None,
        )
        if concept is None:
            pytest.skip("taxonomy has no concept with exactly one required unit")
        env = _Env(taxonomy, config, concept, None, with_unit_map=False)
        assert env.resolver.setUnitForName(env.holder, env.fb) is True
        assert env.fb._aspects["units"] == next(iter(concept.getRequiredUnitQNames()))

    def test_complex_unit_uses_default_currency_denominator(self, taxonomy, config):
        concept = next(
            (
                c
                for c in sorted(taxonomy.concepts, key=str)
                if c.isReportable
                and c.isNumeric
                and c.getRequiredUnitQNames() is None
                and any(
                    unit_id in config.unitIdsToMeasures
                    and unit_id.endswith("_per_Monetary")
                    for unit_id in taxonomy.UTR.getUnitIdsForDataType(c.dataType)
                )
            ),
            None,
        )
        if concept is None:
            pytest.skip("taxonomy has no complex-unit (per-monetary) concept")
        env = _Env(taxonomy, config, concept, None, with_unit_map=False)
        env.report.setDefaultAspect("monetary-units", "EUR")
        assert env.resolver.setUnitForName(env.holder, env.fb) is True
        assert "EUR" in env.fb._aspects["complex-units"]
