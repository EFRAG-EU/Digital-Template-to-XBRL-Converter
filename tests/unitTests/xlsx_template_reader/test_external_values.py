"""External values + partial facts (branch features).

No shipped template uses the ``template_external_values`` named range, so we
synthesize one in-test: take the official 1.3.0 sample, pick a text-block
concept that normally becomes a fact, and add a ``template_external_values``
range naming that concept. The concept should then be reported as a *partial*
fact (its value supplied externally) rather than a normal fact, and
``completePartialFact`` should finalize it.

An externally-valued concept's cell is *expected* to be empty (the value
arrives from an uploaded document), so the partial fact must be registered
whether the cell is populated, empty, or unresolvable.
"""

from collections import defaultdict
from io import BytesIO
from pathlib import Path

import pytest
from markupsafe import Markup
from openpyxl import Workbook
from openpyxl.utils.cell import absolute_coordinate, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from mireport.conversionresults import ConversionResultsBuilder
from mireport.data.disclosures import VSME_DEFAULTS
from mireport.report import InlineReport
from mireport.taxonomy import getTaxonomy, listTaxonomies
from mireport.xlsx_template_reader._binder import WorkbookBinder
from mireport.xlsx_template_reader._bindings import WorkbookBindings
from mireport.xlsx_template_reader._fact_creator import FactCreator
from mireport.xlsx_template_reader._ranges import XbrlConceptCellRangeMetadata
from mireport.xlsx_template_reader._reader import WorkbookReader
from mireport.xlsx_template_reader.processor import XlsxProcessor
from mireport.xlsx_template_reader.util import loadExcelFromPathOrFileLike

SAMPLE_1_3_0 = (
    Path(__file__).parents[3]
    / "digital-templates"
    / "VSME-Digital-Template-Sample-1.3.0.xlsx"
)


def _results() -> ConversionResultsBuilder:
    return ConversionResultsBuilder(consoleOutput=False)


@pytest.fixture(scope="module")
def external_values_case():
    """Return (workbook_bytes, concept) for a synthesized external-values template."""
    assert SAMPLE_1_3_0.is_file(), f"Missing fixture {SAMPLE_1_3_0}"

    # 1. Normal report: find a text-block concept that actually became a fact.
    report = XlsxProcessor.from_file(
        SAMPLE_1_3_0, _results(), VSME_DEFAULTS
    ).createReport()
    chosen = next(
        (f.concept for f in report.facts if f.concept.isTextblock),
        None,
    )
    assert chosen is not None, "1.3.0 sample has no text-block fact to use"

    # 2. Synthesize a workbook with a template_external_values range naming it.
    wb = loadExcelFromPathOrFileLike(SAMPLE_1_3_0)
    ws = wb.worksheets[0]
    scratch = ws.cell(row=1, column=250)  # far-out, otherwise-unused cell
    scratch.value = chosen.qname.localName
    wb.defined_names.add(
        DefinedName(
            "template_external_values",
            attr_text=f"'{ws.title}'!{scratch.coordinate}",
        )
    )
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue(), chosen


@pytest.mark.slow
def test_external_value_becomes_partial_fact(external_values_case):
    blob, concept = external_values_case
    report = XlsxProcessor.from_bytes(blob, _results(), VSME_DEFAULTS).createReport()

    assert report.hasPartialFacts, (
        "Expected a pending partial fact for the external value."
    )
    assert concept in report.partialFactsByConcept
    assert concept not in {f.concept for f in report.facts}, (
        "External-value concept should not also be reported as a normal fact."
    )


@pytest.mark.slow
def test_complete_partial_fact_finalizes_it(external_values_case):
    blob, concept = external_values_case
    report = XlsxProcessor.from_bytes(blob, _results(), VSME_DEFAULTS).createReport()

    report.completePartialFact(concept, Markup("<p>Externally supplied.</p>"))

    assert not report.hasPartialFacts
    assert concept in {f.concept for f in report.facts}


# ---------------------------------------------------------------------------
# Cell-skipping: an empty cell must not prevent the partial fact.
# ---------------------------------------------------------------------------


@pytest.fixture(scope="module")
def taxonomy():
    entry_point = next(ep for ep in listTaxonomies() if "vsme" in ep.lower())
    return getTaxonomy(entry_point)


@pytest.fixture(scope="module")
def external_textblock(taxonomy):
    """A reportable text block whose required dimensions are all defaulted,
    so createSimpleFacts' dimension pre-check doesn't drop it."""
    concept = next(
        (
            c
            for c in sorted(taxonomy.concepts, key=str)
            if c.isTextblock
            and c.isReportable
            and taxonomy.getExplicitDimensionsForPrimaryItem(c)
            <= taxonomy.defaultedDimensions
        ),
        None,
    )
    if concept is None:
        pytest.skip("vsme has no dimension-free reportable text block")
    return concept


def _simple_fact_env(taxonomy, concept, cell_value):
    """A minimal workbook whose only bound named range is `concept` at A1,
    with the concept marked as externally valued."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "S"
    if cell_value is not None:
        ws["A1"] = cell_value
    attr = f"{quote_sheetname('S')}!{absolute_coordinate('A1')}"
    wb.defined_names["ext"] = DefinedName("ext", attr_text=attr)

    results = _results()
    reader = WorkbookReader(wb, results)
    dn = wb.defined_names["ext"]
    crm = reader.peekRange(dn)
    assert crm is not None
    holder = XbrlConceptCellRangeMetadata.fromCellRangeMetadata(crm, concept)
    bindings = WorkbookBindings(
        concept_map={dn: holder},
        tables=[],
        unit_map={},
        preset_dims=defaultdict(dict),
        has_external_value=frozenset({concept}),
        footnote=None,
    )
    report = InlineReport(taxonomy, None)
    creator = FactCreator(bindings, reader, report, results, VSME_DEFAULTS)
    return creator, report


class TestExternalValueCellSkipping:
    def test_empty_cell_still_becomes_partial_fact(self, taxonomy, external_textblock):
        creator, report = _simple_fact_env(taxonomy, external_textblock, None)
        creator.createSimpleFacts()
        assert external_textblock in report.partialFactsByConcept
        assert external_textblock not in {f.concept for f in report.facts}

    def test_populated_cell_registers_partial_fact_and_ignores_value(
        self, taxonomy, external_textblock
    ):
        creator, report = _simple_fact_env(
            taxonomy, external_textblock, "placeholder text"
        )
        creator.createSimpleFacts()
        assert external_textblock in report.partialFactsByConcept
        assert external_textblock not in {f.concept for f in report.facts}


@pytest.fixture(scope="module")
def external_values_case_empty_cell():
    """Like external_values_case, but with the chosen concept's cells blanked —
    the realistic template state when the value comes from a Word document."""
    assert SAMPLE_1_3_0.is_file(), f"Missing fixture {SAMPLE_1_3_0}"

    report = XlsxProcessor.from_file(
        SAMPLE_1_3_0, _results(), VSME_DEFAULTS
    ).createReport()
    chosen = next(
        (f.concept for f in report.facts if f.concept.isTextblock),
        None,
    )
    assert chosen is not None, "1.3.0 sample has no text-block fact to use"

    wb = loadExcelFromPathOrFileLike(SAMPLE_1_3_0)

    # Locate the chosen concept's named range via the binder and blank it.
    binder_results = _results()
    reader = WorkbookReader(wb, binder_results)
    bindings = WorkbookBinder(reader, report.taxonomy, binder_results).bind()
    holder = next(crm for crm in bindings.concept_map.values() if crm.concept == chosen)
    cr = holder.cellRange
    for row in holder.worksheet.iter_rows(
        min_row=cr.min_row, max_row=cr.max_row, min_col=cr.min_col, max_col=cr.max_col
    ):
        for cell in row:
            cell.value = None

    ws = wb.worksheets[0]
    scratch = ws.cell(row=1, column=250)  # far-out, otherwise-unused cell
    scratch.value = chosen.qname.localName
    wb.defined_names.add(
        DefinedName(
            "template_external_values",
            attr_text=f"'{ws.title}'!{scratch.coordinate}",
        )
    )
    buf = BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue(), chosen


@pytest.mark.slow
def test_external_value_with_empty_cell_becomes_partial_fact(
    external_values_case_empty_cell,
):
    blob, concept = external_values_case_empty_cell
    report = XlsxProcessor.from_bytes(blob, _results(), VSME_DEFAULTS).createReport()

    assert report.hasPartialFacts, (
        "An empty cell must not prevent the external-value partial fact."
    )
    assert concept in report.partialFactsByConcept
    assert concept not in {f.concept for f in report.facts}
